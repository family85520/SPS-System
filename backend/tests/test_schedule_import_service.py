from openpyxl import Workbook

from app.services.schedule_import_service import ScheduleImportService


def test_read_header_scans_title_rows_for_chinese_template_headers():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "A部门排班导入模板"
    ws["A3"] = "日期"
    ws["B3"] = "组织ID"
    ws["C3"] = "班次名称"
    ws["D3"] = "值班人员"

    header_row, header_map = ScheduleImportService._read_header(ws)

    assert header_row == 3
    assert header_map == {
        1: "date",
        2: "org_id",
        3: "shift_name",
        4: "members",
    }


def test_read_header_supports_ascii_fallback_headers():
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "date"
    ws["B2"] = "org_id"
    ws["C2"] = "shift_name"
    ws["D2"] = "members"

    header_row, header_map = ScheduleImportService._read_header(ws)

    assert header_row == 2
    assert header_map == {
        1: "date",
        2: "org_id",
        3: "shift_name",
        4: "members",
    }
