"""导出服务 - Excel / PDF 生成"""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, timedelta, datetime
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import OrgStaff
from app.models.export_template import ExportTemplate
from app.models.organization import OrgOrganization
from app.models.schedule import SchSchedule, SchScheduleDetail
from app.models.shift_template import SchShiftTemplate


# ==================== 工具函数 ====================

def _calc_duration(start_time: str, end_time: str) -> float:
    """计算班次时长（小时），跨夜班特殊处理"""
    try:
        sh, sm = map(int, start_time.split(":"))
        eh, em = map(int, end_time.split(":"))
        start_min = sh * 60 + sm
        end_min = eh * 60 + em
        diff = end_min - start_min
        if diff <= 0:
            diff += 24 * 60
        return round(diff / 60, 1)
    except (ValueError, AttributeError):
        return 0.0


WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
STATUS_MAP = {0: "草稿", 1: "已发布", 2: "已撤回", 3: "待审核"}


# ==================== 数据查询 ====================

async def _query_schedule_rows(
    db: AsyncSession,
    start_date: date,
    end_date: date,
    org_id: Optional[int] = None,
) -> list[dict]:
    """查询排班数据，返回按组织/日期排列的行数据"""
    # 查排班记录
    query = (
        select(SchSchedule)
        .where(SchSchedule.date >= start_date, SchSchedule.date <= end_date, SchSchedule.status == 1)
        .order_by(SchSchedule.date, SchSchedule.shift_id)
    )
    if org_id is not None:
        query = query.where(SchSchedule.org_id == org_id)

    schedules = list((await db.execute(query)).scalars().all())
    if not schedules:
        return []

    # 查明细
    schedule_ids = [s.id for s in schedules]
    details = list((await db.execute(
        select(SchScheduleDetail).where(SchScheduleDetail.schedule_id.in_(schedule_ids))
    )).scalars().all())

    detail_map: dict[int, list[SchScheduleDetail]] = defaultdict(list)
    for d in details:
        detail_map[d.schedule_id].append(d)

    # 班次模板
    shift_ids = list({s.shift_id for s in schedules})
    shift_map = {
        s.id: s for s in (await db.execute(
            select(SchShiftTemplate).where(SchShiftTemplate.id.in_(shift_ids))
        )).scalars().all()
    }

    # 人员信息（含旧标签）
    all_staff_ids = list({d.staff_id for d in details})
    staff_rows = (await db.execute(
        select(OrgStaff.id, OrgStaff.name, OrgStaff.employee_no, OrgStaff.tags, OrgStaff.org_id, OrgStaff.phone)
        .where(OrgStaff.id.in_(all_staff_ids))
    )).all()
    staff_info: dict[int, dict] = {}
    for row in staff_rows:
        staff_info[row[0]] = {
            "name": row[1], "employee_no": row[2] or "",
            "tags": row[3] if isinstance(row[3], list) else [],
            "org_id": row[4],
            "phone": row[5] or "",
            "role_tags": [],  # will be populated below
        }

    # 新标识体系：角色标签
    try:
        from app.models import OrgStaffRole, SysRole
        role_rows = (await db.execute(
            select(OrgStaffRole.staff_id, SysRole.name)
            .join(SysRole, OrgStaffRole.role_id == SysRole.id)
            .where(OrgStaffRole.staff_id.in_(all_staff_ids), SysRole.role_type == "tag")
        )).all()
        for staff_id, role_name in role_rows:
            if staff_id in staff_info:
                staff_info[staff_id]["role_tags"].append(role_name)
    except Exception:
        pass

    # 组织名称
    org_ids = list({s.org_id for s in schedules})
    org_map = {
        row[0]: row[1] for row in (await db.execute(
            select(OrgOrganization.id, OrgOrganization.name).where(OrgOrganization.id.in_(org_ids))
        )).all()
    }

    # 组装行数据
    rows = []
    for s in schedules:
        shift = shift_map.get(s.shift_id)
        if not shift:
            continue

        day_details = detail_map.get(s.id, [])
        group_leaders, members, leaders = [], [], []

        # 附带每人完整信息（供模板变量使用）
        member_infos: list[dict] = []

        for d in day_details:
            info = staff_info.get(d.staff_id, {})
            name = info.get("name", f"ID:{d.staff_id}")
            role_tags = info.get("role_tags", [])
            old_tags = info.get("tags", [])

            # 值班领导：使用排班明细的 role_type（来自自动排班引擎）
            is_duty_leader = d.role_type == "leader"
            is_group_leader = "组长" in role_tags

            if is_duty_leader:
                leaders.append(name)
            elif is_group_leader:
                group_leaders.append(name)
            else:
                members.append(name)

            member_infos.append({
                "name": name, "employee_no": info.get("employee_no", ""),
                "phone": info.get("phone", ""), "role_tags": role_tags,
            })

        # 如果没有组长标签，则取第一人当组长（适配无标签场景）
        if not group_leaders and len(members) >= 2:
            group_leaders.append(members.pop(0))

        rows.append({
            "date": str(s.date),
            "weekday": WEEKDAY_NAMES[s.date.isoweekday() - 1],
            "org_name": org_map.get(s.org_id, ""),
            "shift_name": shift.name,
            "time_range": f"{shift.start_time}-{shift.end_time}",
            "group_leaders": group_leaders,
            "members": members,
            "leaders": leaders,
            "member_infos": member_infos,
            "status": STATUS_MAP.get(s.status, "未知"),
            "staff_details": [
                {
                    "staff_id": d.staff_id,
                    "name": staff_info.get(d.staff_id, {}).get("name", ""),
                    "employee_no": staff_info.get(d.staff_id, {}).get("employee_no", ""),
                    "org_id": staff_info.get(d.staff_id, {}).get("org_id"),
                    "org_name": org_map.get(staff_info.get(d.staff_id, {}).get("org_id", 0), ""),
                    "shift_name": shift.name,
                    "date": str(s.date),
                    "duration": _calc_duration(shift.start_time, shift.end_time),
                    "role_tags": staff_info.get(d.staff_id, {}).get("role_tags", []),
                }
                for d in day_details
            ],
        })

    return rows


# ==================== 导出入口 ====================

class ExportService:

    @staticmethod
    async def schedule_pdf(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        org_id: Optional[int] = None,
        unit_name: Optional[str] = None,
        dimension: str = "org",
    ) -> io.BytesIO:
        """生成 PDF 排班表（HTML → PDF，样式与 Excel 一致）"""
        rows = await _query_schedule_rows(db, start_date, end_date, org_id)
        if not rows:
            raise ValueError("所选日期范围内无已发布排班数据")

        if dimension == "person":
            return _build_person_pdf(rows, start_date, end_date, unit_name)
        return _build_org_pdf(rows, unit_name)

    @staticmethod
    async def get_active_shift_names(db: AsyncSession) -> list[str]:
        """获取当前启用的班次名称列表（按开始时间排序）"""
        from app.models.shift_template import SchShiftTemplate
        result = await db.execute(
            select(SchShiftTemplate.name).where(SchShiftTemplate.status == 1).order_by(SchShiftTemplate.start_time)
        )
        return list(result.scalars().all())

    # ==================== 模板管理 ====================

    @staticmethod
    async def list_templates(db: AsyncSession) -> list[dict]:
        result = await db.execute(
            select(ExportTemplate).order_by(ExportTemplate.id.desc())
        )
        return [
            {"id": t.id, "name": t.name, "is_default": t.is_default,
             "description": t.description, "created_at": str(t.created_at)}
            for t in result.scalars().all()
        ]

    @staticmethod
    async def create_template(
        db: AsyncSession, name: str, data: bytes,
        is_default: bool, description: str | None,
    ) -> ExportTemplate:
        if is_default:
            await db.execute(
                __import__('sqlalchemy').update(ExportTemplate).values(is_default=False)
            )
        t = ExportTemplate(name=name, file_data=data, is_default=is_default, description=description)
        db.add(t)
        await db.flush()
        await db.refresh(t)
        return t

    @staticmethod
    async def delete_template(db: AsyncSession, template_id: int):
        t = await db.get(ExportTemplate, template_id)
        if not t:
            raise ValueError("模板不存在")
        await db.delete(t)
        await db.flush()

    @staticmethod
    async def set_default_template(db: AsyncSession, template_id: int):
        t = await db.get(ExportTemplate, template_id)
        if not t:
            raise ValueError("模板不存在")
        from sqlalchemy import update
        await db.execute(update(ExportTemplate).values(is_default=False))
        t.is_default = True
        await db.flush()

    @staticmethod
    async def get_default_template(db: AsyncSession) -> ExportTemplate | None:
        result = await db.execute(
            select(ExportTemplate).where(ExportTemplate.is_default == True).limit(1)
        )
        return result.scalars().first()

    @staticmethod
    async def generate_default_template(db: AsyncSession) -> io.BytesIO:
        """生成系统默认模板文件（含所有占位符，用户可下载后修改）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # 班次数量和列数（从数据库获取启用的班次）
        shift_names = await ExportService.get_active_shift_names(db)
        shift_count = len(shift_names)
        col_count = 8 + (shift_count - 2) * 2 if shift_count > 2 else 8

        wb = Workbook()
        ws = wb.active
        ws.title = "排班表模板"

        hdr_font = Font(name="楷体", bold=True, size=11)
        hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="楷体", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))

        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        c = ws.cell(row=1, column=1, value="{{title}}"); c.font = Font(name="方正小标宋简体", bold=True, size=16)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # 副标题
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        c = ws.cell(row=2, column=1, value="24小时值班电话：                    制表时间：{{start_date}}"); c.font = Font(name="楷体", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # 表头 row 3-4（统一使用 shift_N_* 变量）
        for r in (3, 4):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=r, column=col); cell.font = hdr_font
                cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin
        ws.merge_cells("A3:A4"); ws["A3"] = "日期"
        ws.merge_cells("B3:B4"); ws["B3"] = "星期"
        ws.merge_cells("C3:D3"); ws["C3"] = "{{shift_0_name}}\n（{{shift_0_time}}）"
        ws["C4"] = "值班组长"; ws["D4"] = "值班人员"
        ws.merge_cells("E3:F3"); ws["E3"] = "{{shift_1_name}}\n（{{shift_1_time}}）"
        ws["E4"] = "值班组长"; ws["F4"] = "值班人员"
        if shift_count > 2:
            ws.merge_cells("G3:H3"); ws["G3"] = "{{shift_2_name}}\n（{{shift_2_time}}）"
            ws["G4"] = "组长"; ws["H4"] = "人员"
            ws.merge_cells("I3:I4"); ws["I3"] = "值班领导"
            ws.merge_cells("J3:J4"); ws["J3"] = "岗位职责"
        else:
            ws.merge_cells("G3:G4"); ws["G3"] = "值班领导"
            ws.merge_cells("H3:H4"); ws["H3"] = "岗位职责"

        # 数据行（统一使用 shift_N_leader / shift_N_members）
        data_row = 5
        if shift_count > 2:
            placeholders = [
                "{{date}}", "{{weekday}}",
                "{{shift_0_leader}}", "{{shift_0_members}}",
                "{{shift_1_leader}}", "{{shift_1_members}}",
                "{{shift_2_leader}}", "{{shift_2_members}}",
                "{{duty_leader}}", ""
            ]
            col_count = 10
        else:
            placeholders = [
                "{{date}}", "{{weekday}}",
                "{{shift_0_leader}}", "{{shift_0_members}}",
                "{{shift_1_leader}}", "{{shift_1_members}}",
                "{{duty_leader}}", ""
            ]
            col_count = 8
        for ci, ph in enumerate(placeholders, 1):
            val = "{{#each dates}}" + ph if ci == 1 else ph
            cell = ws.cell(row=data_row, column=ci, value=val)
            cell.font = cell_font; cell.border = thin; cell.alignment = center

        # 职责说明行
        note_row = data_row + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=col_count)
        ws.cell(row=note_row, column=1, value="{{duty_text}}").font = cell_font

        # 页脚
        footer_row = note_row + 1
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=col_count)
        ws.cell(row=footer_row, column=1, value="{{footer}}").font = cell_font

        # 列宽
        for i, w in enumerate([13, 8, 10, 16, 10, 16, 12, 42], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 36; ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 42; ws.row_dimensions[4].height = 28

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return buf

    # ==================== 模板填充引擎 ====================

    @staticmethod
    def fill_template(template_data: bytes, rows: list[dict], start_date: date, end_date: date) -> io.BytesIO:
        """用排班数据填充自定义模板。所有班次统一使用 shift_N_* 变量（N=0,1,2...按开始时间排序）。

        变量清单:
          全局:     {{title}} {{org_name}} {{start_date}} {{end_date}} {{duty_text}} {{footer}}
          每日期:   {{date}} {{weekday}} {{duty_leader}}
          每班次:   {{shift_N_name}} {{shift_N_time}} {{shift_N_leader}} {{shift_N_members}}
          控制:     {{#each dates}} 标记循环行
        """
        from openpyxl import load_workbook
        from copy import copy

        buf = io.BytesIO(template_data)
        wb = load_workbook(buf)
        ws = wb.active

        # ---- 识别所有班次，按开始时间排序决定 shift_0/1/2... ----
        shift_entries: list[tuple[str, str]] = []  # (name, time_range)
        seen = set()
        for rd in rows:
            sn = rd["shift_name"]
            if sn in seen: continue
            seen.add(sn)
            shift_entries.append((sn, rd["time_range"]))
        shift_entries.sort(key=lambda x: x[1])  # 按时段排序
        shift_meta: dict[int, dict] = {}
        for idx, (name, tr) in enumerate(shift_entries):
            shift_meta[idx] = {"name": name, "time_range": tr}

        # ---- 按日分组 ----
        day_map: dict[str, dict] = {}
        shift_personnel: list[set[str]] = [set() for _ in shift_meta]  # 每个班次的人员（用于duty_text）
        for rd in rows:
            d = rd["date"]
            if d not in day_map:
                day_map[d] = {"date": d, "weekday": rd["weekday"], "shifts": {}, "duty_leaders": []}
            shift = rd["shift_name"]
            if shift not in day_map[d]["shifts"]:
                day_map[d]["shifts"][shift] = {"leaders": [], "members": [],
                                                "phones": [], "employee_nos": []}
            day_map[d]["shifts"][shift]["leaders"].extend(rd["group_leaders"])
            day_map[d]["shifts"][shift]["members"].extend(rd["members"])
            # 联系方式 & 工号
            for mi in rd.get("member_infos", []):
                if mi.get("phone"): day_map[d]["shifts"][shift]["phones"].append(mi["phone"])
                if mi.get("employee_no"): day_map[d]["shifts"][shift]["employee_nos"].append(mi["employee_no"])
            # 收集每个班次的人员
            for idx, sm in shift_meta.items():
                if sm["name"] == shift:
                    shift_personnel[idx].update(rd["group_leaders"])
                    shift_personnel[idx].update(rd["members"])
            for name in rd["leaders"]:
                if name not in day_map[d]["duty_leaders"]:
                    day_map[d]["duty_leaders"].append(name)

        sorted_dates = sorted(day_map.keys())
        org_name = rows[0]["org_name"] if rows else ""
        # 收集所有排班人员（用于duty_text排除说明）
        all_personnel: set[str] = set()
        for sp in shift_personnel:
            all_personnel.update(sp)
        admin_names = "、".join(sorted(all_personnel)) if all_personnel else ""

        # ---- 构建单日数据 ----
        def _date_data(d: str) -> dict:
            info = day_map[d]
            data = {"date": d, "weekday": info["weekday"],
                    "duty_leader": "、".join(dict.fromkeys(info["duty_leaders"]))}
            for idx, sm in shift_meta.items():
                sn = sm["name"]
                sd = info["shifts"].get(sn, {"leaders": [], "members": [],
                                               "phones": [], "employee_nos": []})
                # 索引写法：{{shift_0_leader}}
                data[f"shift_{idx}_name"] = sn
                data[f"shift_{idx}_time"] = sm["time_range"]
                data[f"shift_{idx}_leader"] = "、".join(sd["leaders"])
                data[f"shift_{idx}_members"] = "、".join(sd["members"])
                data[f"shift_{idx}_phones"] = "、".join(sd.get("phones", []))
                data[f"shift_{idx}_employee_nos"] = "、".join(sd.get("employee_nos", []))
                # 名称写法：{{白班_leader}}，与索引写法指向同一份数据
                data[f"{sn}_name"] = sn
                data[f"{sn}_time"] = sm["time_range"]
                data[f"{sn}_leader"] = data[f"shift_{idx}_leader"]
                data[f"{sn}_members"] = data[f"shift_{idx}_members"]
                data[f"{sn}_phones"] = data[f"shift_{idx}_phones"]
                data[f"{sn}_employee_nos"] = data[f"shift_{idx}_employee_nos"]
            return data

        # ---- 全局替换 ----
        global_replace = {
            "title": f"{org_name}日常值班工作安排表",
            "org_name": org_name,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "duty_text": f"排班人员：全体监管人员（除{admin_names}）" if admin_names else "",
            "footer": "分管领导：          科室负责人：           制表人：",
        }
        for idx, sm in shift_meta.items():
            global_replace[f"shift_{idx}_name"] = sm["name"]
            global_replace[f"shift_{idx}_time"] = sm["time_range"]
            # 名称别名
            global_replace[f"{sm['name']}_name"] = sm["name"]
            global_replace[f"{sm['name']}_time"] = sm["time_range"]

        def _sub(text: str, data: dict | None = None) -> str:
            if not isinstance(text, str): return text
            for k, v in global_replace.items():
                text = text.replace(f"{{{{{k}}}}}", str(v))
            if data:
                for k, v in data.items():
                    text = text.replace(f"{{{{{k}}}}}", str(v))
            return text

        # ---- 查找 {{#each dates}} 标记行 ----
        each_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{{#each dates}}" in cell.value:
                    each_row = cell.row
                    cell.value = cell.value.replace("{{#each dates}}", "")
                    break
            if each_row is not None:
                break

        if each_row is not None:
            template_row_data = []
            for cell in ws[each_row]:
                val = cell.value
                is_static = isinstance(val, str) and val.startswith("{{#static}}")
                if is_static:
                    val = val[len("{{#static}}"):]
                template_row_data.append({
                    "col": cell.column, "value": val,
                    "font": copy(cell.font), "fill": copy(cell.fill),
                    "border": copy(cell.border), "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "static": is_static,
                })

            n_dates = len(sorted_dates)

            # ---- 保存并解除循环行下方的合并单元格（openpyxl insert_rows 不移合并）----
            saved_merges = []
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row > each_row:
                    saved_merges.append((mc.min_row, mc.max_row, mc.min_col, mc.max_col))
            for min_r, max_r, min_c, max_c in saved_merges:
                ws.unmerge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)

            if n_dates > 1:
                ws.insert_rows(each_row + 1, n_dates - 1)

            # 预处理 {{#static}} 列：用首日数据解析模板变量，全月统一
            static_data = _date_data(sorted_dates[0]) if sorted_dates else {}

            for di, d in enumerate(sorted_dates):
                target_row = each_row + di
                data = _date_data(d)
                for td in template_row_data:
                    c = ws.cell(row=target_row, column=td["col"])
                    c.font = copy(td["font"]); c.fill = copy(td["fill"])
                    c.border = copy(td["border"]); c.alignment = copy(td["alignment"])
                    c.number_format = td["number_format"]
                    if td.get("static"):
                        # 静态列：先用首日数据解析，再叠加上全局替换（跳出 each 循环后统一扫）
                        c.value = _sub(td["value"], static_data)
                    else:
                        c.value = _sub(td["value"], data)

            # ---- 恢复合并单元格（手动计算 insert 后的位置）----
            for min_r, max_r, min_c, max_c in saved_merges:
                adj_min = min_r + n_dates - 1
                adj_max = max_r + n_dates - 1
                try:
                    ws.merge_cells(start_row=adj_min, start_column=min_c,
                                   end_row=adj_max, end_column=max_c)
                except Exception:
                    pass

            # ---- 自动合并：循环区域内相邻同行值相同的单元格 ----
            data_start, data_end = each_row, each_row + n_dates - 1
            cols = {td["col"] for td in template_row_data}
            for col in cols:
                merge_start = data_start
                prev_val = ws.cell(row=data_start, column=col).value
                for r in range(data_start + 1, data_end + 1):
                    cur_val = ws.cell(row=r, column=col).value
                    if cur_val == prev_val and cur_val is not None and cur_val != "":
                        continue  # 延续当前合并组
                    if r - 1 > merge_start:
                        ws.merge_cells(start_row=merge_start, start_column=col,
                                       end_row=r - 1, end_column=col)
                    merge_start = r
                    prev_val = cur_val
                # 最后一组合并
                if data_end > merge_start and prev_val is not None and prev_val != "":
                    ws.merge_cells(start_row=merge_start, start_column=col,
                                   end_row=data_end, end_column=col)
        else:
            # 无 {{#each dates}} 标记 → 单日填充
            default_data = _date_data(sorted_dates[0]) if sorted_dates else {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "{{" in str(cell.value):
                        cell.value = _sub(cell.value, default_data)

        # ---- 全表扫描：替换循环行外的全局变量（{{title}} 等）----
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{{" in str(cell.value):
                    cell.value = _sub(cell.value)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @staticmethod
    async def statistics_pdf(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        org_id: Optional[int] = None,
    ) -> io.BytesIO:
        """PDF 统计报表（HTML → PDF）"""
        rows = await _query_schedule_rows(db, start_date, end_date, org_id)
        if not rows:
            raise ValueError("所选日期范围内无已发布排班数据")
        return _build_stats_pdf(rows, start_date, end_date)

    @staticmethod
    async def statistics_excel(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        org_id: Optional[int] = None,
    ) -> io.BytesIO:
        return await _build_statistics_excel(db, start_date, end_date, org_id)

    @staticmethod
    async def schedule_excel(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        org_id: Optional[int] = None,
        dimension: str = "org",
    ) -> io.BytesIO:
        rows = await _query_schedule_rows(db, start_date, end_date, org_id)
        if not rows:
            raise ValueError("所选日期范围内无已发布排班数据")
        if dimension == "person":
            return _build_person_excel(rows, start_date, end_date)
        # 检查是否有自定义默认模板
        tpl = await ExportService.get_default_template(db)
        if tpl:
            return ExportService.fill_template(tpl.file_data, rows, start_date, end_date)
        return _build_org_excel(rows)


# ==================== Excel 生成 ====================

def _build_org_excel(rows: list[dict]) -> io.BytesIO:
    """按组织维度生成排班表 Excel（每日一行，白班+夜班+行政班并排）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "值班安排表"

    # ===== 样式定义 =====
    title_font = Font(name="方正小标宋简体", bold=True, size=16)
    subtitle_font = Font(name="楷体", size=11)
    hdr_font = Font(name="楷体", bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="楷体", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    footer_font = Font(name="楷体", size=11)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ===== 按日期分组：每天一行 =====
    day_map: dict[str, dict] = {}
    admin_personnel: set[str] = set()

    for rd in rows:
        d = rd["date"]
        if d not in day_map:
            day_map[d] = {
                "date": d, "weekday": rd["weekday"], "org_name": rd["org_name"],
                "day_leaders": [], "day_members": [],
                "night_leaders": [], "night_members": [],
                "duty_leaders": [],  # 仅带班领导（身份标识）
            }
        shift = rd["shift_name"]

        if shift == "行政":
            admin_personnel.update(rd["members"])
            admin_personnel.update(rd["group_leaders"])
        elif shift == "夜班":
            day_map[d]["night_leaders"].extend(rd["group_leaders"])
            day_map[d]["night_members"].extend(rd["members"])
        else:
            day_map[d]["day_leaders"].extend(rd["group_leaders"])
            day_map[d]["day_members"].extend(rd["members"])

        # 值班领导：仅来自身份标识 "领导"
        for name in rd["leaders"]:
            if name not in day_map[d]["duty_leaders"]:
                day_map[d]["duty_leaders"].append(name)

    sorted_dates = sorted(day_map.keys())
    org_name = rows[0]["org_name"] if rows else ""

    # ===== 标题行 =====
    title = f"{org_name}日常值班工作安排表"
    if sorted_dates:
        first = sorted_dates[0]
        last = sorted_dates[-1]
        if first[:7] == last[:7]:
            parts = first[:7].split("-")
            title += f"（{int(parts[0])}年{int(parts[1])}月）"
        else:
            title += f"（{first}~{last}）"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws.cell(row=1, column=1, value=title)
    c.font, c.alignment = title_font, Alignment(horizontal="center", vertical="center")

    # ===== 副标题行 =====
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    c = ws.cell(row=2, column=1,
                value=f"24小时值班电话：                    制表时间：{datetime.now().strftime('%Y年%m月%d日')}")
    c.font = subtitle_font
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ===== 表头（两行） =====
    # 先铺满所有表头格子的样式，再合并（合并后 MergedCell 无法写值但可以提前设样式）
    for r in (3, 4):
        for c in range(1, 9):
            _hdr(ws, r, c, "", hdr_font, hdr_fill, hdr_align, thin_border)

    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    _hdr(ws, 3, 1, "日期", hdr_font, hdr_fill, hdr_align, thin_border)

    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    _hdr(ws, 3, 2, "星期", hdr_font, hdr_fill, hdr_align, thin_border)

    # 提取班次时间（从第一天的数据中获取）
    day_times = ""
    night_times = ""
    for rd in rows:
        if rd["shift_name"] == "白班" and not day_times:
            day_times = rd["time_range"]
        if rd["shift_name"] == "夜班" and not night_times:
            night_times = rd["time_range"]
        if day_times and night_times:
            break

    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    _hdr(ws, 3, 3, f"白班\n（{day_times}）" if day_times else "白班", hdr_font, hdr_fill, hdr_align, thin_border)

    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6)
    _hdr(ws, 3, 5, f"夜班\n（{night_times}）" if night_times else "夜班", hdr_font, hdr_fill, hdr_align, thin_border)

    ws.merge_cells(start_row=3, start_column=7, end_row=4, end_column=7)
    _hdr(ws, 3, 7, "值班领导", hdr_font, hdr_fill, hdr_align, thin_border)

    ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=8)
    _hdr(ws, 3, 8, "岗位职责", hdr_font, hdr_fill, hdr_align, thin_border)

    # 子表头（row 4, cols 3-6 不参与合并）
    _hdr(ws, 4, 3, "值班组长", hdr_font, hdr_fill, hdr_align, thin_border)
    _hdr(ws, 4, 4, "值班人员", hdr_font, hdr_fill, hdr_align, thin_border)
    _hdr(ws, 4, 5, "值班组长", hdr_font, hdr_fill, hdr_align, thin_border)
    _hdr(ws, 4, 6, "值班人员", hdr_font, hdr_fill, hdr_align, thin_border)

    # ===== 数据行 + 值班领导按周合并 =====
    # 先写数据，同时记录每周的起止行和值班领导
    DATA_START = 5
    ri = DATA_START
    week_regions: list[dict] = []  # [{start_row, end_row, leader_text}]

    # 按日期分组为周（周一~周日）
    from itertools import groupby
    def _week_key(d: str):
        dt = date.fromisoformat(d)
        # ISO week: Monday=1 ... Sunday=7. 按周一所属周分组
        iso = dt.isocalendar()
        return (iso[0], iso[1])

    week_groups = []
    for key, group_dates in groupby(sorted_dates, key=_week_key):
        week_groups.append(list(group_dates))

    for week_dates in week_groups:
        week_start_row = ri
        # 整理本周值班领导（取所有天里去重）
        week_leaders: list[str] = []
        for d in week_dates:
            info = day_map[d]
            for name in info["duty_leaders"]:
                if name not in week_leaders:
                    week_leaders.append(name)
        leader_text = "、".join(week_leaders)

        for d in week_dates:
            info = day_map[d]
            vals = [
                d,
                info["weekday"],
                "、".join(info["day_leaders"]) if info["day_leaders"] else "",
                "、".join(info["day_members"]),
                "、".join(info["night_leaders"]) if info["night_leaders"] else "",
                "、".join(info["night_members"]),
                leader_text,  # 每周相同
                "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = cell_font
                c.border = thin_border
                # 日期、星期、组长、人员、值班领导 居中；岗位职责 左对齐
                c.alignment = left_wrap if ci == 8 else center
            ri += 1

        week_end_row = ri - 1
        if week_end_row > week_start_row:
            ws.merge_cells(start_row=week_start_row, start_column=7,
                           end_row=week_end_row, end_column=7)
        week_regions.append({
            "start_row": week_start_row,
            "end_row": week_end_row,
            "leader_text": leader_text,
        })

    # ===== 岗位职责（合并） =====
    admin_names = "、".join(sorted(admin_personnel)) if admin_personnel else ""
    duty_text = (
        "一、工作职责：\n"
        "（一）开展应急值守，监测、记录煤矿瓦斯超限、安全事故等异常情况，"
        "落实上级交办工作任务，并及时向带班领导报告，按要求妥善做好。\n"
        "（二）开展网上巡查，巡查煤矿视频监控规范性以及安全生产违法违规行为，"
        "巡查安全监测系统、人员定位系统规范性，并做好记录。\n"
        "（三）加强安全监管，根据工作需要，每月至少安排2个巡查片区辅助开展日常安全监管及现场核查工作。\n\n"
        "二、工作要求：\n"
        "（一）实行白班、夜班两班倒机制，白班：8:30-19:00，夜班：19:00-次日8:30，"
        "每天安排2名值班人员共同值班，每天由当班组长统筹安排值班工作，"
        "遇异常情况及时汇报给带班领导，按局机关信息报送有关规定进行处置。"
        "周末及节假日由一人到监控中心值班，另一人机动备勤。\n"
        "（二）按时接班，确因特殊原因确实不能按时接班，需提前与对方做好沟通，严禁出现空岗。\n"
        "（三）严格值班纪律，根据需要做好调度工作，严禁出现离岗、脱岗、睡岗。\n\n"
        f"    排班人员及负责人：全体监管人员（除{admin_names}）"
    )
    last_data_row = ri - 1
    ws.merge_cells(start_row=DATA_START, start_column=8, end_row=last_data_row, end_column=8)
    c = ws.cell(row=5, column=8, value=duty_text)
    c.font = cell_font
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ===== 页脚 =====
    footer_row = last_data_row + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    c = ws.cell(row=footer_row, column=1,
                value="分管领导：                    科室负责人：                     制表人：")
    c.font = footer_font
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ===== 列宽 ====
    col_widths = [13, 8, 10, 16, 10, 16, 12, 42]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行高
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 42  # 班次时间双行
    ws.row_dimensions[4].height = 28
    for r in range(DATA_START, last_data_row + 1):
        ws.row_dimensions[r].height = 36

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _hdr(ws, row, col, value, font, fill, align, border):
    """Helper: set header cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.alignment, c.border = font, fill, align, border


def _build_person_excel(rows: list[dict], start_date: date, end_date: date) -> io.BytesIO:
    """按人员维度生成排班表 Excel（透视表）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 构建人员维度数据
    staff_map: dict[int, dict] = {}
    for row in rows:
        for sd in row.get("staff_details", []):
            sid = sd["staff_id"]
            if sid not in staff_map:
                staff_map[sid] = {
                    "name": sd["name"],
                    "employee_no": sd["employee_no"],
                    "org_name": sd["org_name"],
                    "dates": {},
                    "total_shifts": 0,
                    "total_hours": 0.0,
                }
            s = staff_map[sid]
            s["dates"][sd["date"]] = sd["shift_name"]
            s["total_shifts"] += 1
            s["total_hours"] += sd["duration"]

    # 生成日期列
    date_cols = []
    cur = start_date
    while cur <= end_date:
        date_cols.append({"date": str(cur), "label": f"{cur.month}/{cur.day}({WEEKDAY_NAMES[cur.isoweekday()-1]})"})
        cur += timedelta(days=1)

    # 排序
    staff_list = sorted(staff_map.values(), key=lambda x: x["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "按人员排班"

    # 样式
    hdr_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="0A63D8", end_color="0A63D8", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="E6EAF0"),
        right=Side(style="thin", color="E6EAF0"),
        top=Side(style="thin", color="E6EAF0"),
        bottom=Side(style="thin", color="E6EAF0"),
    )
    even_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    no_shift_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

    # 表头
    fixed_headers = ["姓名", "工号", "组织"]
    tail_headers = ["总班次", "总工时(h)"]
    all_headers = fixed_headers + [d["label"] for d in date_cols] + tail_headers
    for col, h in enumerate(all_headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, thin

    # 数据行
    for ri, staff in enumerate(staff_list, 2):
        ws.cell(row=ri, column=1, value=staff["name"]).font = Font(size=10, bold=True)
        ws.cell(row=ri, column=1, value=staff["name"]).border = thin
        ws.cell(row=ri, column=1, value=staff["name"]).alignment = center
        ws.cell(row=ri, column=2, value=staff["employee_no"]).font = cell_font
        ws.cell(row=ri, column=2, value=staff["employee_no"]).border = thin
        ws.cell(row=ri, column=2, value=staff["employee_no"]).alignment = center
        ws.cell(row=ri, column=3, value=staff["org_name"]).font = cell_font
        ws.cell(row=ri, column=3, value=staff["org_name"]).border = thin

        for di, dc in enumerate(date_cols):
            col_idx = len(fixed_headers) + di + 1
            shift_name = staff["dates"].get(dc["date"], "")
            c = ws.cell(row=ri, column=col_idx, value=shift_name)
            c.font = cell_font; c.alignment = center; c.border = thin
            c.fill = PatternFill(start_color="EBF5FF", end_color="EBF5FF", fill_type="solid") if shift_name else no_shift_fill

        tail_start = len(fixed_headers) + len(date_cols) + 1
        ws.cell(row=ri, column=tail_start, value=staff["total_shifts"]).font = cell_font
        ws.cell(row=ri, column=tail_start, value=staff["total_shifts"]).border = thin
        ws.cell(row=ri, column=tail_start, value=staff["total_shifts"]).alignment = center
        ws.cell(row=ri, column=tail_start + 1, value=staff["total_hours"]).font = cell_font
        ws.cell(row=ri, column=tail_start + 1, value=staff["total_hours"]).border = thin
        ws.cell(row=ri, column=tail_start + 1, value=staff["total_hours"]).alignment = center

        if ri % 2 == 0:
            for ci in range(1, len(all_headers) + 1):
                c = ws.cell(row=ri, column=ci)
                if not c.fill or c.fill.start_color.index in ("00000000", "FAFAFA", "EBF5FF"):
                    c.fill = even_fill

    # 列宽
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    for i in range(len(date_cols)):
        ws.column_dimensions[get_column_letter(len(fixed_headers) + i + 1)].width = 10
    ws.column_dimensions[get_column_letter(tail_start)].width = 10
    ws.column_dimensions[get_column_letter(tail_start + 1)].width = 12
    ws.freeze_panes = "D2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==================== 统计报表 Excel ====================

async def _build_statistics_excel(
    db: AsyncSession,
    start_date: date,
    end_date: date,
    org_id: Optional[int] = None,
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = await _query_schedule_rows(db, start_date, end_date, org_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "排班统计"

    hdr_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="0A63D8", end_color="0A63D8", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="E6EAF0"), right=Side(style="thin", color="E6EAF0"),
        top=Side(style="thin", color="E6EAF0"), bottom=Side(style="thin", color="E6EAF0"),
    )
    even_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    # 人员路径统计
    staff_stats: dict[int, dict] = {}
    for row in rows:
        for sd in row.get("staff_details", []):
            sid = sd["staff_id"]
            if sid not in staff_stats:
                staff_stats[sid] = {
                    "name": sd["name"], "employee_no": sd["employee_no"],
                    "org_name": sd["org_name"], "total": 0, "total_hours": 0.0,
                    "night": 0, "weekend": 0, "holiday": 0,
                }
            s = staff_stats[sid]
            s["total"] += 1
            s["total_hours"] += sd["duration"]
            shift_name = sd["shift_name"]
            if shift_name == "夜班":
                s["night"] += 1
            dt = date.fromisoformat(sd["date"])
            if dt.isoweekday() >= 6:
                s["weekend"] += 1

    headers = ["排名", "姓名", "工号", "组织", "排班天数", "总工时(h)", "夜班次数", "周末次数"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, thin

    sorted_stats = sorted(staff_stats.values(), key=lambda x: -x["total"])
    for ri, s in enumerate(sorted_stats, 2):
        vals = [ri - 1, s["name"], s["employee_no"], s["org_name"],
                s["total"], s["total_hours"], s["night"], s["weekend"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = cell_font, thin
            c.alignment = center
            if ri % 2 == 0:
                c.fill = even_fill

    col_widths = [8, 12, 12, 16, 12, 14, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==================== PDF 生成（ReportLab） ====================

def _register_cn_font():
    """注册中文字体，返回字体名。"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    paths = [
        ("C:/Windows/Fonts/simsun.ttc", "SimSun"),
        ("C:/Windows/Fonts/kaiu.ttf", "KaiTi"),
        ("C:/Windows/Fonts/msyh.ttc", "MSYH"),
    ]
    for fp, name in paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(name, fp))
                return name
            except Exception:
                continue
    return "Helvetica"


def _build_org_pdf(rows: list[dict], unit_name: str = "") -> io.BytesIO:
    """按组织维度生成 PDF（reportlab，样式与 Excel 一致）"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    P = Paragraph
    cn = _register_cn_font()
    s_title = ParagraphStyle('T', fontName=cn, fontSize=14, leading=20, alignment=TA_CENTER)
    s_cell = ParagraphStyle('C', fontName=cn, fontSize=7, leading=10, alignment=TA_CENTER)
    s_left = ParagraphStyle('L', fontName=cn, fontSize=7, leading=10, alignment=TA_LEFT)

    # 按日分组（复用 Excel 逻辑）
    day_map: dict[str, dict] = {}
    admin_personnel: set[str] = set()
    for rd in rows:
        d = rd["date"]
        if d not in day_map:
            day_map[d] = {"date": d, "weekday": rd["weekday"],
                           "day_leaders": [], "day_members": [],
                           "night_leaders": [], "night_members": [],
                           "duty_leaders": []}
        shift = rd["shift_name"]
        if shift == "行政":
            admin_personnel.update(rd["members"]); admin_personnel.update(rd["group_leaders"])
        elif shift == "夜班":
            day_map[d]["night_leaders"].extend(rd["group_leaders"])
            day_map[d]["night_members"].extend(rd["members"])
        else:
            day_map[d]["day_leaders"].extend(rd["group_leaders"])
            day_map[d]["day_members"].extend(rd["members"])
        for name in rd["leaders"]:
            if name not in day_map[d]["duty_leaders"]:
                day_map[d]["duty_leaders"].append(name)

    sorted_dates = sorted(day_map.keys())
    org_name = rows[0]["org_name"] if rows else ""
    title = f"{org_name}日常值班工作安排表"
    if sorted_dates and sorted_dates[0][:7] == sorted_dates[-1][:7]:
        parts = sorted_dates[0][:7].split("-")
        title += f"（{int(parts[0])}年{int(parts[1])}月）"

    day_time, night_time = "", ""
    for rd in rows:
        if rd["shift_name"] == "白班" and not day_time: day_time = rd["time_range"]
        if rd["shift_name"] == "夜班" and not night_time: night_time = rd["time_range"]

    admin_names = "、".join(sorted(admin_personnel)) if admin_personnel else ""

    # 构建表格数据
    hdr = [
        P("日期", s_cell), P("星期", s_cell),
        P(f"白班\n({day_time})" if day_time else "白班", s_cell), P("", s_cell),
        P(f"夜班\n({night_time})" if night_time else "夜班", s_cell), P("", s_cell),
        P("值班领导", s_cell), P("岗位职责", s_cell),
    ]
    hdr2 = [
        P("", s_cell), P("", s_cell),
        P("值班组长", s_cell), P("值班人员", s_cell),
        P("值班组长", s_cell), P("值班人员", s_cell),
        P("", s_cell), P("", s_cell),
    ]
    table_data = [hdr, hdr2]

    for d in sorted_dates:
        info = day_map[d]
        table_data.append([
            P(d, s_cell), P(info["weekday"], s_cell),
            P("、".join(info["day_leaders"]) if info["day_leaders"] else "", s_cell),
            P("、".join(info["day_members"]), s_cell),
            P("、".join(info["night_leaders"]) if info["night_leaders"] else "", s_cell),
            P("、".join(info["night_members"]), s_cell),
            P("、".join(dict.fromkeys(info["duty_leaders"])), s_cell),
            P("", s_cell),
        ])

    buf = io.BytesIO()
    pw = landscape(A4)[0] - 20*mm
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    elements = [Paragraph(title, s_title), Spacer(1, 3*mm)]

    col_w = [pw*0.08, pw*0.06, pw*0.10, pw*0.15, pw*0.10, pw*0.15, pw*0.10, pw*0.26]
    t = Table(table_data, colWidths=col_w, repeatRows=2)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), cn),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 1), colors.Color(0.85, 0.89, 0.95)),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('SPAN', (2, 0), (3, 0)), ('SPAN', (4, 0), (5, 0)),
    ]))
    elements.append(t)

    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph(
        f"排班人员及负责人：全体监管人员（除{admin_names}）", s_left))
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(
        "分管领导：                    科室负责人：                     制表人：", s_left))

    doc.build(elements)
    buf.seek(0)
    return buf


def _build_person_pdf(rows: list[dict], start_date: date, end_date: date, unit_name: str = "") -> io.BytesIO:
    """按人员维度生成 PDF"""
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER

    P = Paragraph
    cn = _register_cn_font()
    s_cell = ParagraphStyle('C', fontName=cn, fontSize=5, leading=7, alignment=TA_CENTER)
    s_hdr = ParagraphStyle('H', fontName=cn, fontSize=6, leading=8, alignment=TA_CENTER)

    # 人员维度数据
    staff_map: dict[int, dict] = {}
    for row in rows:
        for sd in row.get("staff_details", []):
            sid = sd["staff_id"]
            if sid not in staff_map:
                staff_map[sid] = {"name": sd["name"], "employee_no": sd["employee_no"],
                                   "org_name": sd["org_name"], "dates": {},
                                   "total_shifts": 0, "total_hours": 0.0}
            s = staff_map[sid]
            s["dates"][sd["date"]] = sd["shift_name"]
            s["total_shifts"] += 1; s["total_hours"] += sd["duration"]

    date_cols = []
    cur = start_date
    while cur <= end_date:
        date_cols.append({"date": str(cur), "label": f"{cur.month}/{cur.day}"})
        cur += timedelta(days=1)

    staff_list = sorted(staff_map.values(), key=lambda x: x["name"])

    header = [P("姓名", s_hdr), P("工号", s_hdr), P("组织", s_hdr)]
    header += [P(d["label"], s_hdr) for d in date_cols]
    header += [P("总班次", s_hdr), P("总工时", s_hdr)]
    table_data = [header]

    for s in staff_list:
        row = [P(s["name"], s_cell), P(s["employee_no"], s_cell), P(s["org_name"], s_cell)]
        for dc in date_cols:
            row.append(P(s["dates"].get(dc["date"], ""), s_cell))
        row.append(P(str(s["total_shifts"]), s_cell))
        row.append(P(str(s["total_hours"]), s_cell))
        table_data.append(row)

    buf = io.BytesIO()
    ncols = len(header)
    col_w = [36, 36, 50] + [22] * len(date_cols) + [28, 32]
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                            leftMargin=6*mm, rightMargin=6*mm,
                            topMargin=8*mm, bottomMargin=8*mm)
    t = Table(table_data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), cn),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.04, 0.39, 0.85)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.Color(0.9, 0.9, 0.9)),
    ]))
    doc.build([t])
    buf.seek(0)
    return buf


def _build_stats_pdf(rows: list[dict], start_date: date, end_date: date) -> io.BytesIO:
    """统计报表 PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER

    P = Paragraph
    cn = _register_cn_font()
    s = ParagraphStyle('C', fontName=cn, fontSize=8, leading=11, alignment=TA_CENTER)

    staff_stats: dict[int, dict] = {}
    for row in rows:
        for sd in row.get("staff_details", []):
            sid = sd["staff_id"]
            if sid not in staff_stats:
                staff_stats[sid] = {"name": sd["name"], "employee_no": sd["employee_no"],
                                     "org_name": sd["org_name"], "total": 0, "total_hours": 0.0,
                                     "night": 0, "weekend": 0}
            st = staff_stats[sid]
            st["total"] += 1; st["total_hours"] += sd["duration"]
            if sd["shift_name"] == "夜班": st["night"] += 1
            if date.fromisoformat(sd["date"]).isoweekday() >= 6: st["weekend"] += 1

    sorted_stats = sorted(staff_stats.values(), key=lambda x: -x["total"])
    header = [P(h, s) for h in ["排名","姓名","工号","组织","排班天数","总工时(h)","夜班","周末"]]
    table_data = [header]
    for i, st in enumerate(sorted_stats, 1):
        table_data.append([P(str(v), s) for v in [i, st["name"], st["employee_no"],
                           st["org_name"], st["total"], st["total_hours"], st["night"], st["weekend"]]])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    t = Table(table_data, colWidths=[24, 48, 48, 60, 48, 48, 36, 36], repeatRows=1)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), cn),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.04, 0.39, 0.85)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
    ]))
    doc.build([t])
    buf.seek(0)
    return buf
