"""Excel template download and import for manual schedules."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime

from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import OrgOrganization, OrgStaff
from app.models.schedule import SchSchedule, SchScheduleDetail
from app.models.shift_template import SchShiftTemplate
from app.services.schedule_service import _sync_pairing_from_schedule


TEMPLATE_HEADERS = [
    "日期",
    "组织ID",
    "组织名称",
    "班次ID",
    "班次名称",
    "值班人员",
    "值班领导",
    "备注",
]

HEADER_ALIASES = {
    "date": "date",
    "schedule_date": "date",
    "日期": "date",
    "排班日期": "date",
    "org_id": "org_id",
    "organization_id": "org_id",
    "组织ID": "org_id",
    "部门ID": "org_id",
    "org_name": "org_name",
    "organization_name": "org_name",
    "组织名称": "org_name",
    "部门名称": "org_name",
    "shift_id": "shift_id",
    "shift_template_id": "shift_id",
    "班次ID": "shift_id",
    "班次模板ID": "shift_id",
    "shift_name": "shift_name",
    "班次名称": "shift_name",
    "班次": "shift_name",
    "members": "members",
    "member_names": "members",
    "值班人员": "members",
    "人员": "members",
    "成员": "members",
    "leaders": "leaders",
    "leader_names": "leaders",
    "值班领导": "leaders",
    "领导": "leaders",
    "note": "note",
    "备注": "note",
}

SPLIT_PATTERN = re.compile(r"[,\u3001;；，\n\r]+")


@dataclass
class ImportRow:
    row_number: int
    schedule_date: date
    org_id: int
    shift_id: int
    member_ids: list[int] = field(default_factory=list)
    leader_ids: list[int] = field(default_factory=list)
    note: str | None = None


class ScheduleImportService:
    """Import manually maintained schedule rows from the standard Excel file."""

    @staticmethod
    async def generate_template(
        db: AsyncSession,
        org_id: int | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> io.BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = "排班导入"
        dict_ws = wb.create_sheet("字典")
        guide_ws = wb.create_sheet("填写说明")

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        required_font = Font(color="C00000", bold=True)
        thin = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, header in enumerate(TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        for col, width in enumerate([14, 10, 18, 10, 16, 34, 28, 30], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

        shifts_query = select(SchShiftTemplate).where(SchShiftTemplate.status == 1)
        if org_id is not None:
            shifts_query = shifts_query.where(
                (SchShiftTemplate.org_id == org_id) | (SchShiftTemplate.org_id.is_(None))
            )
        shifts = list((await db.execute(
            shifts_query.order_by(SchShiftTemplate.start_time, SchShiftTemplate.id)
        )).scalars().all())

        orgs_query = select(OrgOrganization).where(OrgOrganization.status == 1)
        if org_id is not None:
            orgs_query = orgs_query.where(OrgOrganization.id == org_id)
        orgs = list((await db.execute(
            orgs_query.order_by(OrgOrganization.sort_order, OrgOrganization.id)
        )).scalars().all())

        staff_query = select(OrgStaff).where(OrgStaff.status == 1)
        if org_id is not None:
            staff_query = staff_query.where(OrgStaff.org_id == org_id)
        staff_list = list((await db.execute(
            staff_query.order_by(OrgStaff.org_id, OrgStaff.id)
        )).scalars().all())

        dict_ws.append(["组织ID", "组织名称"])
        for org in orgs:
            dict_ws.append([org.id, org.name])
        dict_ws.append([])
        shift_start_row = dict_ws.max_row + 1
        dict_ws.append(["班次ID", "班次名称", "开始时间", "结束时间"])
        for shift in shifts:
            dict_ws.append([shift.id, shift.name, shift.start_time, shift.end_time])
        dict_ws.append([])
        staff_start_row = dict_ws.max_row + 1
        dict_ws.append(["人员ID", "姓名", "工号", "组织ID"])
        for staff in staff_list:
            dict_ws.append([staff.id, staff.name, staff.employee_no, staff.org_id])

        for col in range(1, 5):
            dict_ws.column_dimensions[get_column_letter(col)].width = 18

        guide_lines = [
            "1. 红色字段建议必填：日期、组织ID或组织名称、班次ID或班次名称、值班人员/值班领导至少一项。",
            "2. 值班人员和值班领导支持用逗号、顿号、分号或换行分隔。",
            "3. 人员可填写姓名、工号或人员ID；同一组织内姓名重复时请改填工号或人员ID。",
            "4. 导入会创建草稿排班；若同日期/组织/班次已有草稿或已撤回记录，会覆盖人员明细。",
            "5. 若同日期/组织/班次已有已发布或待审核记录，系统会拒绝导入，避免覆盖正式排班。",
        ]
        for idx, line in enumerate(guide_lines, 1):
            guide_ws.cell(row=idx, column=1, value=line)
        guide_ws.column_dimensions["A"].width = 110

        data_start = 2
        data_end = 300
        if start_date and end_date and orgs and shifts:
            row = data_start
            current = start_date
            while current <= end_date and row <= data_end:
                for shift in shifts:
                    if row > data_end:
                        break
                    org = orgs[0]
                    ws.cell(row=row, column=1, value=current)
                    ws.cell(row=row, column=2, value=org.id)
                    ws.cell(row=row, column=3, value=org.name)
                    ws.cell(row=row, column=4, value=shift.id)
                    ws.cell(row=row, column=5, value=shift.name)
                    row += 1
                current = date.fromordinal(current.toordinal() + 1)
        else:
            for row in range(data_start, min(data_end, 20) + 1):
                ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"

        for row in range(data_start, data_end + 1):
            ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
            for col in range(1, len(TEMPLATE_HEADERS) + 1):
                ws.cell(row=row, column=col).alignment = Alignment(
                    vertical="center", wrap_text=True
                )
                ws.cell(row=row, column=col).border = thin
            for col in (1, 2, 4, 5):
                ws.cell(row=row, column=col).font = required_font

        if len(orgs) > 0:
            org_name_range = f"'字典'!$B$2:$B${1 + len(orgs)}"
            dv = DataValidation(type="list", formula1=org_name_range, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"C{data_start}:C{data_end}")
        if len(shifts) > 0:
            shift_name_range = (
                f"'字典'!$B${shift_start_row + 1}:$B${shift_start_row + len(shifts)}"
            )
            dv = DataValidation(type="list", formula1=shift_name_range, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"E{data_start}:E{data_end}")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    async def import_excel(
        db: AsyncSession,
        file_data: bytes,
        default_org_id: int | None = None,
    ) -> dict:
        rows = await ScheduleImportService._parse_rows(db, file_data, default_org_id)
        if not rows:
            raise ValueError("未读取到可导入的排班数据")

        duplicate_keys: dict[tuple[date, int, int], int] = {}
        for row in rows:
            key = (row.schedule_date, row.org_id, row.shift_id)
            if key in duplicate_keys:
                raise ValueError(
                    f"第 {row.row_number} 行与第 {duplicate_keys[key]} 行日期/组织/班次重复"
                )
            duplicate_keys[key] = row.row_number

        existing = list((await db.execute(
            select(SchSchedule).where(
                SchSchedule.date.in_([row.schedule_date for row in rows]),
                SchSchedule.org_id.in_([row.org_id for row in rows]),
                SchSchedule.shift_id.in_([row.shift_id for row in rows]),
            )
        )).scalars().all())
        existing_by_key = {
            (schedule.date, schedule.org_id, schedule.shift_id): schedule
            for schedule in existing
        }

        locked = [
            schedule for schedule in existing
            if schedule.status in SchSchedule.LOCKED_STATUSES
            and (schedule.date, schedule.org_id, schedule.shift_id) in duplicate_keys
        ]
        if locked:
            first = locked[0]
            raise ValueError(
                f"{first.date} 组织ID:{first.org_id} 班次ID:{first.shift_id} "
                "已发布或待审核，不能通过导入覆盖"
            )

        created_count = 0
        updated_count = 0
        detail_count = 0
        touched_schedules: list[SchSchedule] = []

        for row in rows:
            key = (row.schedule_date, row.org_id, row.shift_id)
            schedule = existing_by_key.get(key)
            if schedule:
                await db.execute(
                    delete(SchScheduleDetail).where(SchScheduleDetail.schedule_id == schedule.id)
                )
                schedule.leader_staff_id = row.leader_ids[0] if row.leader_ids else None
                schedule.status = SchSchedule.STATUS_DRAFT
                schedule.source = "import"
                updated_count += 1
            else:
                schedule = SchSchedule(
                    date=row.schedule_date,
                    shift_id=row.shift_id,
                    org_id=row.org_id,
                    leader_staff_id=row.leader_ids[0] if row.leader_ids else None,
                    status=SchSchedule.STATUS_DRAFT,
                    source="import",
                )
                db.add(schedule)
                await db.flush()
                existing_by_key[key] = schedule
                created_count += 1

            leader_set = set(row.leader_ids)
            ordered_staff_ids = list(dict.fromkeys(row.leader_ids + row.member_ids))
            for staff_id in ordered_staff_ids:
                db.add(SchScheduleDetail(
                    schedule_id=schedule.id,
                    staff_id=staff_id,
                    role_type="leader" if staff_id in leader_set else "member",
                    is_substitute=False,
                    note=row.note,
                ))
                detail_count += 1
            touched_schedules.append(schedule)

        await db.flush()
        for schedule in touched_schedules:
            await _sync_pairing_from_schedule(db, schedule)

        return {
            "created_count": created_count,
            "updated_count": updated_count,
            "detail_count": detail_count,
            "total_count": len(rows),
            "message": (
                f"导入完成：新增 {created_count} 条，覆盖 {updated_count} 条，"
                f"人员明细 {detail_count} 条"
            ),
        }

    @staticmethod
    async def _parse_rows(
        db: AsyncSession,
        file_data: bytes,
        default_org_id: int | None,
    ) -> list[ImportRow]:
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
        except Exception as exc:
            raise ValueError(f"Excel 文件读取失败：{exc}") from exc

        ws = wb["排班导入"] if "排班导入" in wb.sheetnames else wb.active
        header_row, header_map = ScheduleImportService._read_header(ws)
        header_fields = set(header_map.values())
        has_shift = bool({"shift_id", "shift_name"} & header_fields)
        has_staff = bool({"members", "leaders"} & header_fields)
        if "date" not in header_fields or not has_shift or not has_staff:
            raise ValueError(
                "模板表头不正确：至少需要包含日期、班次ID或班次名称、"
                "值班人员或值班领导。请下载系统标准排班导入模板后填写"
            )

        orgs = list((await db.execute(
            select(OrgOrganization).where(OrgOrganization.status == 1)
        )).scalars().all())
        shifts = list((await db.execute(
            select(SchShiftTemplate).where(SchShiftTemplate.status == 1)
        )).scalars().all())
        staff_list = list((await db.execute(
            select(OrgStaff).where(OrgStaff.status == 1)
        )).scalars().all())

        org_by_id = {org.id: org for org in orgs}
        org_by_name: dict[str, list[OrgOrganization]] = {}
        for org in orgs:
            org_by_name.setdefault(_norm(org.name), []).append(org)

        shift_by_id = {shift.id: shift for shift in shifts}
        shift_by_name: dict[str, list[SchShiftTemplate]] = {}
        for shift in shifts:
            shift_by_name.setdefault(_norm(shift.name), []).append(shift)

        staff_by_id = {staff.id: staff for staff in staff_list}
        staff_by_no = {
            _norm(staff.employee_no): staff
            for staff in staff_list
            if staff.employee_no
        }
        staff_by_org_name: dict[tuple[int, str], list[OrgStaff]] = {}
        for staff in staff_list:
            staff_by_org_name.setdefault((staff.org_id, _norm(staff.name)), []).append(staff)

        rows: list[ImportRow] = []
        for row_number in range(header_row + 1, ws.max_row + 1):
            row_values = {
                key: _cell_value(ws.cell(row=row_number, column=col).value)
                for col, key in header_map.items()
            }
            if not any(row_values.values()):
                continue

            schedule_date = _parse_date_value(row_values.get("date"), row_number)
            org_id = _resolve_org_id(
                row_values.get("org_id"),
                row_values.get("org_name"),
                default_org_id,
                org_by_id,
                org_by_name,
                row_number,
            )
            shift_id = _resolve_shift_id(
                row_values.get("shift_id"),
                row_values.get("shift_name"),
                org_id,
                shift_by_id,
                shift_by_name,
                row_number,
            )
            leader_ids = _resolve_staff_tokens(
                row_values.get("leaders"),
                org_id,
                staff_by_id,
                staff_by_no,
                staff_by_org_name,
                row_number,
                "值班领导",
            )
            member_ids = _resolve_staff_tokens(
                row_values.get("members"),
                org_id,
                staff_by_id,
                staff_by_no,
                staff_by_org_name,
                row_number,
                "值班人员",
            )
            if not leader_ids and not member_ids:
                raise ValueError(f"第 {row_number} 行值班人员和值班领导不能同时为空")

            rows.append(ImportRow(
                row_number=row_number,
                schedule_date=schedule_date,
                org_id=org_id,
                shift_id=shift_id,
                member_ids=[sid for sid in member_ids if sid not in set(leader_ids)],
                leader_ids=leader_ids,
                note=row_values.get("note") or None,
            ))

        return rows

    @staticmethod
    def _read_header(ws) -> tuple[int, dict[int, str]]:
        best_row = 1
        best_map: dict[int, str] = {}
        for row in range(1, min(ws.max_row, 10) + 1):
            current: dict[int, str] = {}
            for col in range(1, ws.max_column + 1):
                raw = _cell_value(ws.cell(row=row, column=col).value)
                if not raw:
                    continue
                key = _header_key(raw)
                if key:
                    current[col] = key
            current_fields = set(current.values())
            if (
                "date" in current_fields
                and {"shift_id", "shift_name"} & current_fields
                and {"members", "leaders"} & current_fields
            ):
                return row, current
            if len(current) > len(best_map):
                best_row = row
                best_map = current
        return best_row, best_map


def _cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _norm(value: str | None) -> str:
    return (value or "").strip().lower()


def _header_key(value: str | None) -> str | None:
    text = (value or "").strip()
    if not text:
        return None
    if text in HEADER_ALIASES:
        return HEADER_ALIASES[text]
    compact = re.sub(r"[\s:：\r\n]+", "", text)
    if compact in HEADER_ALIASES:
        return HEADER_ALIASES[compact]
    for alias, key in HEADER_ALIASES.items():
        if alias in compact:
            return key
    return None


def _parse_date_value(value: str | None, row_number: int) -> date:
    if not value:
        raise ValueError(f"第 {row_number} 行日期不能为空")
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"第 {row_number} 行日期格式错误，应为 YYYY-MM-DD") from exc


def _parse_int(value: str | None) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if not text.isdigit():
        return None
    return int(text)


def _resolve_org_id(
    org_id_raw: str | None,
    org_name: str | None,
    default_org_id: int | None,
    org_by_id: dict[int, OrgOrganization],
    org_by_name: dict[str, list[OrgOrganization]],
    row_number: int,
) -> int:
    org_id = _parse_int(org_id_raw) or default_org_id
    if org_id is not None:
        if org_id not in org_by_id:
            raise ValueError(f"第 {row_number} 行组织ID不存在或已停用：{org_id}")
        return org_id

    name = _norm(org_name)
    matches = org_by_name.get(name, [])
    if not name:
        raise ValueError(f"第 {row_number} 行组织ID或组织名称必须填写")
    if len(matches) != 1:
        raise ValueError(f"第 {row_number} 行组织名称无法唯一匹配：{org_name}")
    return matches[0].id


def _resolve_shift_id(
    shift_id_raw: str | None,
    shift_name: str | None,
    org_id: int,
    shift_by_id: dict[int, SchShiftTemplate],
    shift_by_name: dict[str, list[SchShiftTemplate]],
    row_number: int,
) -> int:
    shift_id = _parse_int(shift_id_raw)
    if shift_id is not None:
        shift = shift_by_id.get(shift_id)
        if not shift:
            raise ValueError(f"第 {row_number} 行班次ID不存在或已停用：{shift_id}")
        if shift.org_id not in (None, org_id):
            raise ValueError(f"第 {row_number} 行班次ID不属于该组织：{shift_id}")
        return shift_id

    name = _norm(shift_name)
    if not name:
        raise ValueError(f"第 {row_number} 行班次ID或班次名称必须填写")
    matches = [
        shift for shift in shift_by_name.get(name, [])
        if shift.org_id in (None, org_id)
    ]
    if len(matches) != 1:
        raise ValueError(f"第 {row_number} 行班次名称无法唯一匹配：{shift_name}")
    return matches[0].id


def _resolve_staff_tokens(
    raw: str | None,
    org_id: int,
    staff_by_id: dict[int, OrgStaff],
    staff_by_no: dict[str, OrgStaff],
    staff_by_org_name: dict[tuple[int, str], list[OrgStaff]],
    row_number: int,
    field_name: str,
) -> list[int]:
    if not raw:
        return []
    result: list[int] = []
    tokens = [item.strip() for item in SPLIT_PATTERN.split(str(raw)) if item.strip()]
    for token in tokens:
        staff = None
        sid = _parse_int(token)
        if sid is not None:
            staff = staff_by_id.get(sid)
        if staff is None:
            staff = staff_by_no.get(_norm(token))
        if staff is None:
            matches = staff_by_org_name.get((org_id, _norm(token)), [])
            if len(matches) == 1:
                staff = matches[0]
            elif len(matches) > 1:
                raise ValueError(
                    f"第 {row_number} 行{field_name}姓名重复，请填写工号或人员ID：{token}"
                )
        if staff is None:
            raise ValueError(f"第 {row_number} 行{field_name}无法匹配人员：{token}")
        if staff.org_id != org_id:
            raise ValueError(f"第 {row_number} 行{field_name}不属于当前组织：{token}")
        if staff.id not in result:
            result.append(staff.id)
    return result
